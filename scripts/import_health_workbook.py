"""Import a verified health workbook without storing source data in Git."""

from __future__ import annotations

import argparse
import hashlib
import json
import sys
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable
from openpyxl import load_workbook


SCHEMA_VERSION = "health-workbook-v1"
METRIC_MAP = {
    "ActiveEnergyBurned": "activity.active_energy",
    "AppleExerciseTime": "activity.exercise_minutes",
    "AppleSleepingBreathingDisturbances": "sleep.breathing_disturbances",
    "AppleSleepingWristTemperature": "sleep.wrist_temperature",
    "AppleStandHour": "activity.stand_hours",
    "AppleStandTime": "activity.stand_minutes",
    "AppleWalkingSteadiness": "mobility.walking_steadiness",
    "BasalEnergyBurned": "activity.basal_energy",
    "BodyMass": "body.weight",
    "DistanceWalkingRunning": "activity.walk_run_distance",
    "DistanceCycling": "activity.cycling_distance",
    "EnvironmentalAudioExposure": "audio.environmental_exposure",
    "EnvironmentalAudioExposureEvent": "audio.environmental_exposure_event",
    "FlightsClimbed": "activity.flights_climbed",
    "Handwashing": "hygiene.handwashing",
    "HandwashingEvent": "hygiene.handwashing",
    "HeadphoneAudioExposure": "audio.headphone_exposure",
    "HeadphoneAudioExposureEvent": "audio.headphone_exposure_event",
    "HeartRate": "heart.rate",
    "HeartRateRecoveryOneMinute": "heart.recovery_one_minute",
    "HeartRateVariabilitySDNN": "heart.hrv_sdnn",
    "Height": "body.height",
    "OxygenSaturation": "respiratory.spo2",
    "PhysicalEffort": "activity.physical_effort",
    "RespiratoryRate": "respiratory.rate",
    "RestingHeartRate": "heart.resting_rate",
    "RunningGroundContactTime": "running.ground_contact_time",
    "RunningPower": "running.power",
    "RunningSpeed": "running.speed",
    "RunningStrideLength": "running.stride_length",
    "RunningVerticalOscillation": "running.vertical_oscillation",
    "SixMinuteWalkTestDistance": "mobility.six_minute_walk_distance",
    "SleepAnalysis": "sleep.analysis",
    "SleepDurationGoal": "sleep.duration_goal",
    "StairAscentSpeed": "mobility.stair_ascent_speed",
    "StairDescentSpeed": "mobility.stair_descent_speed",
    "StepCount": "activity.steps",
    "TimeInDaylight": "environment.daylight_minutes",
    "VO2Max": "fitness.vo2max",
    "WalkingAsymmetryPercentage": "mobility.walking_asymmetry",
    "WalkingDoubleSupportPercentage": "mobility.walking_double_support",
    "WalkingHeartRateAverage": "heart.walking_average",
    "WalkingSpeed": "mobility.walking_speed",
    "WalkingStepLength": "mobility.walking_step_length",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Prepare a Personal OS health import package")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output", type=Path, required=True, help="Write an ignored local import package for the PWA")
    parser.add_argument("--upstream-apple-export", type=Path, help="Original Apple Health XML used to build the workbook")
    parser.add_argument("--upstream-export-date", help="Export timestamp reported by Apple Health")
    parser.add_argument("--upstream-record-count", type=int, help="Top-level Record count verified in the Apple export")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if not args.workbook.is_file():
        raise SystemExit("Workbook not found")
    artifact_digest = sha256_file(args.workbook)
    if args.upstream_apple_export and not args.upstream_apple_export.is_file():
        raise SystemExit("Upstream Apple Health export not found")
    source_digest = sha256_file(args.upstream_apple_export) if args.upstream_apple_export else artifact_digest
    source_type = "apple_health_export" if args.upstream_apple_export else "health_workbook"
    source_name = "Export Apple Salute riconciliato" if args.upstream_apple_export else args.workbook.name
    workbook = load_workbook(args.workbook, read_only=True, data_only=True)
    required = {"Raw_Metriche", "Peso_e_Misure", "Sonno", "Allenamenti"}
    if not required.issubset(set(workbook.sheetnames)):
        raise SystemExit("Workbook does not match the expected health schema")

    daily = list(parse_daily_metrics(workbook["Raw_Metriche"]))
    excluded_weight_count = count_manual_or_estimated_weights(workbook["Peso_e_Misure"])
    sleep = list(parse_sleep(workbook["Sonno"]))
    workouts = list(parse_workouts(workbook["Allenamenti"]))
    counts = {"daily_metrics": len(daily), "sleep": len(sleep), "workouts": len(workouts)}
    args.output.parent.mkdir(parents=True, exist_ok=True)
    package = {
        "schema_version": SCHEMA_VERSION,
        "source_type": source_type,
        "source_name": source_name,
        "source_sha256": source_digest,
        "artifact_name": args.workbook.name,
        "artifact_sha256": artifact_digest,
        "exported_at": args.upstream_export_date,
        "upstream_record_count": args.upstream_record_count,
        "coverage_start": min((row["observed_on"] for row in daily), default=None),
        "coverage_end": max((row["observed_on"] for row in daily), default=None),
        "transformation": "apple-health-daily-aggregation-v1" if args.upstream_apple_export else "health-workbook-v1",
        "validation": "reconciled" if args.upstream_apple_export else "pending",
        "import_mode": "snapshot",
        "row_counts": counts,
        "excluded_counts": {"manual_or_estimated_weights": excluded_weight_count},
        "daily_metrics": daily,
        "sleep": sleep,
        "workouts": workouts,
    }
    args.output.write_text(json.dumps(serialize(package), ensure_ascii=True, separators=(",", ":")), encoding="utf-8")
    print(json.dumps({
        "status": "staged_package",
        "canonical": False,
        "output": str(args.output.resolve()),
        "source_sha256": source_digest,
        "artifact_sha256": artifact_digest,
        "counts": counts,
        "excluded_counts": {"manual_or_estimated_weights": excluded_weight_count},
    }, ensure_ascii=True))
    return 0


def parse_daily_metrics(sheet: Any) -> Iterable[dict[str, Any]]:
    for row in sheet.iter_rows(min_row=5, values_only=True):
        observed, metric, source, unit, count, value_sum, avg, minimum, maximum, first, last = row[:11]
        if not observed or not metric or not source or not unit:
            continue
        yield compact({
            "observed_on": iso_date(observed),
            "metric_key": METRIC_MAP.get(str(metric), metric_key(str(metric))),
            "source_label": str(source)[:160],
            "unit": str(unit)[:40],
            "record_count": count,
            "value_sum": value_sum,
            "value_avg": avg,
            "value_min": minimum,
            "value_max": maximum,
            "value_first": first,
            "value_last": last,
        })


def count_manual_or_estimated_weights(sheet: Any) -> int:
    count = 0
    for row in sheet.iter_rows(min_row=5, values_only=True):
        observed, value = row[:2]
        if observed and value is not None:
            count += 1
    return count


def parse_sleep(sheet: Any) -> Iterable[dict[str, Any]]:
    for row in sheet.iter_rows(min_row=5, values_only=True):
        observed, detected, valid, efficiency, core, deep, rem, awake, status = row[:9]
        if not observed:
            continue
        yield compact({
            "observed_on": iso_date(observed), "detected_hours": detected, "valid_hours": valid,
            "efficiency": efficiency, "core_minutes": core, "deep_minutes": deep,
            "rem_minutes": rem, "awake_minutes": awake, "source_status": status,
        })


def parse_workouts(sheet: Any) -> Iterable[dict[str, Any]]:
    for source_row, row in enumerate(sheet.iter_rows(min_row=13, values_only=True), 13):
        observed, activity, duration, distance, energy, avg_hr, max_hr, speed, route, source = row[:10]
        if not observed or not activity:
            continue
        yield compact({
            "observed_on": iso_date(observed), "activity_type": str(activity)[:80],
            "duration_minutes": duration, "distance_km": distance, "energy_kcal": energy,
            "average_heart_rate": avg_hr, "maximum_heart_rate": max_hr,
            "running_speed_kmh": speed, "route_file_name": route, "source_label": source,
            "source_row": source_row,
        })


def metric_key(value: str) -> str:
    output = []
    for index, char in enumerate(value):
        if char.isupper() and index and output[-1] != ".":
            output.append("_")
        output.append(char.lower() if char.isalnum() else "_")
    return "health." + "".join(output).strip("_")[:100]


def compact(row: dict[str, Any]) -> dict[str, Any]:
    return {key: value for key, value in row.items() if value is not None and value != ""}


def iso_date(value: Any) -> str:
    if isinstance(value, (date, datetime)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    return str(value)[:10]


def serialize(value: Any) -> Any:
    if isinstance(value, dict):
        return {key: serialize(item) for key, item in value.items()}
    if isinstance(value, list):
        return [serialize(item) for item in value]
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while chunk := handle.read(1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


if __name__ == "__main__":
    sys.exit(main())
